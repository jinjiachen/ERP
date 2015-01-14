#coding=cp936
from openpyxl import Workbook
from openpyxl import load_workbook
import volume_calc

def exc(x1,cho1,cho2,m,x2,x6,length1,width1,height1,length2,width2,height2,length3,width3,height3,length4,width4,height4):
    wb=load_workbook('D:\sample.xlsx')
    ws=wb.active
    ws.title=u'���۵�'
#����
    ws.merge_cells('A1:F1')
    ws.cell(row=1,column=1).value=u'ѹ���������۱�'
    ws.cell(row=2,column=1).value=u'���'
    ws.cell(row=2,column=2).value=u'����'
    ws.cell(row=2,column=3).value=u'����'
    ws.cell(row=2,column=4).value=u'����'
    ws.cell(row=2,column=5).value=u'����'
    ws.cell(row=2,column=6).value=u'����'
#���п�ʼд������
    #Ͳ��
    ws.cell(row=3,column=1).value=1
    ws.cell(row=3,column=2).value=u'Ͳ��'
    ws.cell(row=3,column=3).value=x1
    ws.cell(row=3,column=4).value=1
    ws.cell(row=3,column=5).value=m
    ws.cell(row=3,column=6).value=ws.cell(row=3,column=5).value*ws.cell(row=3,column=4).value
    #��ͷ
    ws.cell(row=4,column=1).value=2
    ws.cell(row=4,column=2).value=u'��ͷ'
    ws.cell(row=4,column=3).value='Q245R'
    ws.cell(row=4,column=4).value=2
    ws.cell(row=4,column=5).value=x2
    ws.cell(row=4,column=6).value=ws.cell(row=4,column=5).value*ws.cell(row=4,column=4).value
    #�ӹ�
    if cho1.encode('cp936')<>'��':
        i=6
        ws.cell(row=5,column=1).value=3
        ws.cell(row=5,column=2).value=u'�ӹ�'
        ws.cell(row=5,column=3).value=x6
        ws.cell(row=5,column=4).value=1
        ws.cell(row=5,column=5).value=u'����'
        ws.cell(row=5,column=6).value=ws.cell(row=5,column=5).value*ws.cell(row=5,column=4).value
    else:
        i=5

    if cho2==1:
    #����
        ws.cell(row=i,column=1).value=i-2
        ws.cell(row=i,column=2).value=u'����'
        ws.cell(row=i,column=3).value='Q245B'
        ws.cell(row=i,column=4).value=1
        ws.cell(row=i,column=5).value=plate(length1,width1,height1)
        ws.cell(row=i,column=6).value=ws.cell(row=i,column=5).value*ws.cell(row=i,column=4).value
    #���
        ws.cell(row=i+1,column=1).value=i-1
        ws.cell(row=i+1,column=2).value=u'���'
        ws.cell(row=i+1,column=3).value='Q245B'
        ws.cell(row=i+1,column=4).value=1
        ws.cell(row=i+1,column=5).value=plate(length2,width2,height2)
        ws.cell(row=i+1,column=6).value=ws.cell(row=i+1,column=5).value*ws.cell(row=i+1,column=4).value
    #�װ�
        ws.cell(row=i+2,column=1).value=i
        ws.cell(row=i+2,column=2).value=u'�װ�'
        ws.cell(row=i+2,column=3).value='Q245B'
        ws.cell(row=i+2,column=4).value=1
        ws.cell(row=i+2,column=5).value=plate(length3,width3,height3)
        ws.cell(row=i+2,column=6).value=ws.cell(row=i+2,column=5).value*ws.cell(row=i+2,column=4).value
    #���
        ws.cell(row=i+3,column=1).value=i+1
        ws.cell(row=i+3,column=2).value=u'���'
        ws.cell(row=i+3,column=3).value='Q245B'
        ws.cell(row=i+3,column=4).value=1
        ws.cell(row=i+3,column=5).value=plate(length4,width4,height4)
        ws.cell(row=i+3,column=6).value=ws.cell(row=i+3,column=5).value*ws.cell(row=i+3,column=4).value
    elif cho2==2:
        #����
        ws.cell(row=i,column=1).value=4
        ws.cell(row=i,column=2).value=u'����'
        ws.cell(row=i,column=3).value='Q245B'
        ws.cell(row=i,column=4).value=1
        ws.cell(row=i,column=5).value=plate(length1,width1,height1)
        ws.cell(row=i,column=6).value=ws.cell(row=i,column=5).value*ws.cell(row=i,column=4).value
    #���
        ws.cell(row=i+1,column=1).value=5
        ws.cell(row=i+1,column=2).value=u'���'
        ws.cell(row=i+1,column=3).value='Q245B'
        ws.cell(row=i+1,column=4).value=1
        ws.cell(row=i+1,column=5).value=plate(length2,width2,height2)
        ws.cell(row=i+1,column=6).value=ws.cell(row=i+1,column=5).value*ws.cell(row=i+1,column=4).value
    #�װ�
        ws.cell(row=i+2,column=1).value=6
        ws.cell(row=i+2,column=2).value=u'�װ�'
        ws.cell(row=i+2,column=3).value='Q245B'
        ws.cell(row=i+2,column=4).value=1
        ws.cell(row=i+2,column=5).value=plate(length3,width3,height3)
        ws.cell(row=i+2,column=6).value=ws.cell(row=i+2,column=5).value*ws.cell(row=i+2,column=4).value
    

    wb.save('D:\output.xlsx')


def plate(L,W,H):#����ƽ��
    L=float(L)
    W=float(W)
    H=float(H)
    m=L*W*H/1000**3*7850
    return(m)



