#coding=cp936
from openpyxl import Workbook
from openpyxl import load_workbook
import volume_calc

def exc(x1,x2,x3,x4,x5,x6,x7,i):
   wb=load_workbook('D:\LCC.xlsx')
   ws=wb.active
   ws.cell(row=i,column=2).value=x1
   ws.cell(row=i,column=3).value=x2
   ws.cell(row=i,column=5).value=x3
   ws.cell(row=i,column=6).value=x4
   ws.cell(row=i,column=7).value=x5+u'×'+x6+u'×'+x7
   


   wb.save('D:\sample'+'s.xlsx')



